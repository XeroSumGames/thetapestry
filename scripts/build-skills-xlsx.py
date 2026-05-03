"""Generate xse-skills-catalog.xlsx from the SRD §05 / Appendix B canonical
catalog (lib/xse-schema.ts). Two sheets: full skill list + by-attribute roll-up.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# -------- Data (matches lib/xse-schema.ts SKILLS verbatim) --------
SKILLS = [
    ('Animal Handling',    'INF', False, 'Understanding how to work with animals, from basic obedience to herd management'),
    ('Athletics',          'PHY', False, 'Fitness, agility, stamina, and coordination, including climbing, jumping, swimming, and overcoming obstacles'),
    ('Barter',             'INF', False, 'Arranging deals, enticing buyers, appraising goods, haggling for the best outcome, and closing deals'),
    ('Demolitions*',       'PHY', True,  'The manufacture and use of explosives, ranging from improvised charges to precision military demolitions'),
    ('Driving',            'DEX', False, 'Drive any vehicle with confidence and finesse, this allows for reckless maneuvers without wrecking'),
    ('Entertainment',      'INF', False, 'The charisma and talent to captivate an audience through music, song, acting, comedy, storytelling, or other form of performance'),
    ('Farming',            'ACU', False, 'Knowing how to grow crops or raise livestock at scale to sustain large groups of people'),
    ('Gambling',           'ACU', False, 'The understanding of underlying mechanics behind games of chance, risk, and reward, and the confidence of knowing when to bet or fold'),
    ('Heavy Weapons*',     'PHY', True,  'The operation of complex, large-scale battlefield weapons like machine guns, launchers, and artillery'),
    ('Inspiration',        'INF', False, 'Being able to boost the morale of individuals or groups or motivate them behind a shared vision or belief'),
    ('Lock-Picking*',      'ACU', True,  'Bypassing locks and security devices to open them without keys or codes'),
    ('Manipulation',       'INF', False, 'Getting others to think, believe, or act in ways that they may not have otherwise done'),
    ('Mechanic*',          'RSN', True,  'Diagnose, repair, maintain, or build complex machines, tools, vehicles, and systems'),
    ('Medicine*',          'RSN', True,  'Providing first aid, diagnosis, treatment, emergency stabilization and advanced medical care to the injured or ill'),
    ('Melee Combat',       'PHY', False, 'Training with melee weapons to improve close-quarters precision, accuracy and damage'),
    ('Navigation',         'ACU', False, 'Innately able to discern directions, remember routes, and plot accurate courses'),
    ('Psychology*',        'RSN', True,  'Leveraging an understanding of human behavior to influence, predict, exploit, or manipulate outcomes'),
    ('Ranged Combat',      'DEX', False, 'Accurately and safely using projectile weapons, ranging from thrown objects to sniper rifles'),
    ('Research',           'RSN', False, 'Being able to efficiently organize, distill, and absorb information to quickly become well informed on any subject'),
    ('Scavenging',         'ACU', False, 'Finding and evaluating missed, hidden, or discarded items that still have use for survival or trade'),
    ('Sleight of Hand',    'DEX', False, 'Well practiced in performing sleight-of-hand tricks, palming, pickpocketing, concealment, and creating subtle diversions'),
    ('Specific Knowledge', 'RSN', False, 'Knowledge about the history, layout, and secrets of a specific area, community, person, or discipline'),
    ('Stealth',            'PHY', False, 'Avoid notice, moving unseen, sticking to the shadows, and avoiding detection'),
    ('Streetwise',         'ACU', False, 'Instinctively being able to navigate urban environments, read situations for danger, and identify underworld resources'),
    ('Survival',           'ACU', False, 'Knowing how to survive in the wild, live off the land, and track people or animals'),
    ('Tactics*',           'RSN', True,  'The application of battlefield or interpersonal strategies in order to gain a situational advantage or upper hand'),
    ('Tinkerer',           'DEX', False, 'Being adept at fixing, modifying, or improving machines, gear, or weapons as well as the ability to improvise inventions'),
    ('Unarmed Combat',     'PHY', False, 'Knowledge and practice of grappling, fist fight, bare fists or martial arts, and body control'),
    ('Weaponsmith*',       'DEX', True,  'Crafting, repairing, and modifying weapons to ensure reliability and effectiveness'),
]

BY_ATTR = [
    ('RSN (Reason)',      6, 'Mechanic*, Medicine*, Psychology*, Research, Specific Knowledge, Tactics*'),
    ('ACU (Acuity)',      7, 'Farming, Gambling, Lock-Picking*, Navigation, Scavenging, Streetwise, Survival'),
    ('PHY (Physicality)', 6, 'Athletics, Demolitions*, Heavy Weapons*, Melee Combat, Stealth, Unarmed Combat'),
    ('INF (Influence)',   5, 'Animal Handling, Barter, Entertainment, Inspiration, Manipulation'),
    ('DEX (Dexterity)',   5, 'Driving, Ranged Combat, Sleight of Hand, Tinkerer, Weaponsmith*'),
]

# -------- Styling primitives --------
HEADER_FILL  = PatternFill('solid', start_color='1F2937')   # dark slate
HEADER_FONT  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
HEADER_ALIGN = Alignment(horizontal='center', vertical='center')

BODY_FONT      = Font(name='Arial', size=11)
BODY_BOLD_FONT = Font(name='Arial', size=11, bold=True)
VOC_FILL       = PatternFill('solid', start_color='E0F2FF')  # subtle blue highlight
ATTR_COLORS = {
    'RSN': '4F46E5',  # indigo
    'ACU': '0EA5E9',  # sky
    'PHY': 'DC2626',  # red
    'INF': 'F59E0B',  # amber
    'DEX': '10B981',  # emerald
}

THIN = Side(style='thin', color='D4D4D4')
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

# -------- Build workbook --------
wb = Workbook()

# Sheet 1: Skills
ws1 = wb.active
ws1.title = 'Skills'
ws1.sheet_properties.tabColor = 'C0392B'  # red — matches the SRD's accent

headers = ['#', 'Skill', 'Attr', 'Voc.', 'Description']
ws1.append(headers)
for col_idx, _ in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col_idx)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGN
    cell.border = BORDER

for i, (name, attr, voc, desc) in enumerate(SKILLS, start=1):
    row = i + 1  # 1 = header
    ws1.cell(row=row, column=1, value=i)
    ws1.cell(row=row, column=2, value=name)
    ws1.cell(row=row, column=3, value=attr)
    ws1.cell(row=row, column=4, value='★' if voc else '')
    ws1.cell(row=row, column=5, value=desc)

    # Apply uniform styling to body cells, with vocational highlight
    fill = VOC_FILL if voc else None
    for col_idx in range(1, 6):
        c = ws1.cell(row=row, column=col_idx)
        c.font = BODY_BOLD_FONT if col_idx == 2 else BODY_FONT
        c.border = BORDER
        if fill:
            c.fill = fill
    # Center #, Attr, Voc.
    ws1.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws1.cell(row=row, column=3).alignment = Alignment(horizontal='center', vertical='center')
    ws1.cell(row=row, column=4).alignment = Alignment(horizontal='center', vertical='center')
    ws1.cell(row=row, column=5).alignment = Alignment(wrap_text=True, vertical='top')

    # Color the attribute cell text
    attr_color = ATTR_COLORS.get(attr)
    if attr_color:
        ws1.cell(row=row, column=3).font = Font(name='Arial', size=11, bold=True, color=attr_color)

# Column widths
widths = {'A': 5, 'B': 22, 'C': 8, 'D': 7, 'E': 80}
for col, w in widths.items():
    ws1.column_dimensions[col].width = w

# Row heights — let description wrap nicely
ws1.row_dimensions[1].height = 24
for r in range(2, len(SKILLS) + 2):
    ws1.row_dimensions[r].height = 30

ws1.freeze_panes = 'A2'

# -------- Sheet 2: By Attribute --------
ws2 = wb.create_sheet('By Attribute')
ws2.sheet_properties.tabColor = '1A4FA5'  # blue

headers2 = ['Attribute', 'Count', 'Skills']
ws2.append(headers2)
for col_idx, _ in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col_idx)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGN
    cell.border = BORDER

for i, (attr, count, skill_list) in enumerate(BY_ATTR, start=2):
    ws2.cell(row=i, column=1, value=attr)
    ws2.cell(row=i, column=2, value=count)
    ws2.cell(row=i, column=3, value=skill_list)
    for col_idx in range(1, 4):
        c = ws2.cell(row=i, column=col_idx)
        c.font = BODY_BOLD_FONT if col_idx == 1 else BODY_FONT
        c.border = BORDER
    ws2.cell(row=i, column=2).alignment = Alignment(horizontal='center', vertical='center')
    ws2.cell(row=i, column=3).alignment = Alignment(wrap_text=True, vertical='top')

    # Color the attribute text
    short = attr.split()[0]  # 'RSN' from 'RSN (Reason)'
    color = ATTR_COLORS.get(short)
    if color:
        ws2.cell(row=i, column=1).font = Font(name='Arial', size=11, bold=True, color=color)

# Total row
total_row = len(BY_ATTR) + 2
ws2.cell(row=total_row, column=1, value='Total')
ws2.cell(row=total_row, column=2, value=f'=SUM(B2:B{total_row - 1})')
ws2.cell(row=total_row, column=3, value='')
for col_idx in range(1, 4):
    c = ws2.cell(row=total_row, column=col_idx)
    c.font = Font(name='Arial', size=11, bold=True)
    c.fill = PatternFill('solid', start_color='F3F4F6')
    c.border = BORDER
ws2.cell(row=total_row, column=2).alignment = Alignment(horizontal='center', vertical='center')

# Column widths
widths2 = {'A': 22, 'B': 10, 'C': 80}
for col, w in widths2.items():
    ws2.column_dimensions[col].width = w

ws2.row_dimensions[1].height = 24
for r in range(2, total_row + 1):
    ws2.row_dimensions[r].height = 30

ws2.freeze_panes = 'A2'

# -------- Save --------
import os
out = r'C:\TheTapestry\docs\Rules\xse-skills-catalog.xlsx'
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print(f'Wrote: {out}')
